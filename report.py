import logging
import importlib.util
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill


REQUIRED_COLUMNS = ["date", "product", "quantity", "unit_price"]
COLUMN_ALIASES = {
    "date": "date",
    "日付": "date",
    "売上日": "date",
    "product": "product",
    "商品": "product",
    "商品名": "product",
    "quantity": "quantity",
    "数量": "quantity",
    "unit_price": "unit_price",
    "単価": "unit_price",
    "category": "category",
    "カテゴリ": "category",
    "カテゴリー": "category",
}
GROUP_BY_COLUMNS = {"product": "商品別", "category": "カテゴリ別"}
DETAIL_COLUMN_LABELS = {
    "date": "日付",
    "product": "商品",
    "category": "カテゴリ",
    "quantity": "数量",
    "unit_price": "単価",
    "source_file": "元ファイル",
    "source_row": "元CSV行",
    "amount": "金額",
}
SUMMARY_COLUMN_LABELS = {
    "rank": "順位",
    "product": "商品",
    "category": "カテゴリ",
    "total_amount": "売上合計",
    "transaction_count": "件数",
    "total_quantity": "数量合計",
    "average_unit_price": "平均単価",
    "total_amount_ratio": "売上構成比",
}
MONTHLY_TREND_COLUMN_LABELS = {
    "month": "月",
    "total_quantity": "数量合計",
    "total_amount": "金額合計",
}
DAILY_TREND_COLUMN_LABELS = {
    "date": "日付",
    "total_amount": "売上合計",
    "transaction_count": "件数",
    "total_quantity": "数量合計",
    "average_unit_price": "平均単価",
}
CSV_ENCODINGS = ["utf-8-sig", "utf-8", "cp932"]
SUPPORTED_INPUT_SUFFIXES = {".csv", ".xlsx", ".xls"}
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
TOTAL_FILL = PatternFill(fill_type="solid", fgColor="EAF4E2")
TITLE_FONT = Font(bold=True, size=14)
HEADER_FONT = Font(bold=True)
TOTAL_FONT = Font(bold=True)
NUMBER_FORMAT = "#,##0"
PERCENT_FORMAT = "0.0%"
DEFAULT_STYLE_CONFIG = {
    "header_fill": "D9EAF7",
    "total_fill": "EAF4E2",
    "title_size": 14,
    "chart_height": 7,
    "chart_width": 14,
}


def merge_style_config(style_config: dict | None) -> dict:
    merged = DEFAULT_STYLE_CONFIG.copy()
    if style_config:
        merged.update(style_config)
    return merged


def merge_column_aliases(column_aliases: dict | None) -> dict:
    merged = COLUMN_ALIASES.copy()
    if column_aliases:
        for source_column, target_column in column_aliases.items():
            merged[str(source_column).strip()] = str(target_column).strip()
    return merged


@dataclass(frozen=True)
class ValidationIssue:
    issue: str
    message: str
    source_file: str
    source_row: int | None


class DataValidationError(ValueError):
    def __init__(self, issues: list[ValidationIssue]) -> None:
        self.issues = issues
        super().__init__("\n".join(format_validation_issues(issues)))


def format_validation_issues(issues: list[ValidationIssue], limit: int = 20) -> list[str]:
    lines = []
    for issue in issues[:limit]:
        location = issue.source_file
        if issue.source_row is not None:
            location = f"{location}:{issue.source_row}"
        lines.append(f"{issue.message} 該当行: {location}")
    extra_count = max(len(issues) - limit, 0)
    if extra_count:
        lines.append(f"ほか{extra_count}件のエラーがあります。")
    return lines


def write_validation_error_report(error: DataValidationError, output_file: Path) -> Path:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    rows = [
        {
            "issue": issue.issue,
            "message": issue.message,
            "source_file": issue.source_file,
            "source_row": issue.source_row,
        }
        for issue in error.issues
    ]
    pd.DataFrame(rows).to_csv(output_file, index=False, encoding="utf-8-sig")
    return output_file


def create_validation_error_rows(error: DataValidationError | None = None) -> pd.DataFrame:
    columns = ["source_file", "source_row", "message", "date", "product", "category", "quantity", "unit_price", "amount"]
    if error is None:
        return pd.DataFrame(columns=columns)
    rows = [
        {
            "source_file": issue.source_file,
            "source_row": issue.source_row,
            "message": issue.message,
            "date": "",
            "product": "",
            "category": "",
            "quantity": "",
            "unit_price": "",
            "amount": "",
        }
        for issue in error.issues
    ]
    return pd.DataFrame(rows, columns=columns)


def read_csv_with_fallback(file: Path) -> pd.DataFrame:
    last_error: Exception | None = None
    for encoding in CSV_ENCODINGS:
        try:
            return pd.read_csv(file, encoding=encoding)
        except UnicodeDecodeError as exc:
            last_error = exc
    raise UnicodeDecodeError(
        "csv",
        b"",
        0,
        1,
        f"{file.name} を対応エンコーディング（UTF-8 または CP932）で読み込めませんでした。",
    ) from last_error


def _expand_input_patterns(pattern: str) -> tuple[str, ...]:
    suffix = Path(pattern).suffix.lower()
    if suffix in SUPPORTED_INPUT_SUFFIXES:
        return tuple(f"{pattern[: -len(suffix)]}{input_suffix}" for input_suffix in sorted(SUPPORTED_INPUT_SUFFIXES))
    return (pattern,)


def find_sales_files(input_path: Path, pattern: str = "*.csv") -> list[Path]:
    if not input_path.exists():
        raise FileNotFoundError(f"入力フォルダが見つかりません: {input_path}")

    if input_path.is_file():
        if input_path.suffix.lower() not in SUPPORTED_INPUT_SUFFIXES:
            raise ValueError(
                f"対応していない入力ファイル形式です: {input_path.name}。"
                "対応形式は .csv / .xlsx / .xls です。"
            )
        return [input_path]

    if not input_path.is_dir():
        raise ValueError(f"入力パスがファイルまたはフォルダではありません: {input_path}")

    files_by_path: dict[Path, Path] = {}
    for expanded_pattern in _expand_input_patterns(pattern):
        for file in input_path.glob(expanded_pattern):
            if (
                file.is_file()
                and file.suffix.lower() in SUPPORTED_INPUT_SUFFIXES
                and not file.name.startswith("sample_")
            ):
                files_by_path[file.resolve()] = file
    input_files = sorted(files_by_path.values(), key=lambda path: path.name)
    if not input_files:
        raise FileNotFoundError(
            "読み込み対象の売上ファイルが見つかりません。"
            f"入力フォルダ: {input_path}。"
            f"ファイル名パターン: {pattern}。"
            "対応形式は .csv / .xlsx / .xls です。"
            "sample_ で始まるファイルはサンプルとして除外されます。"
        )
    return input_files


def read_excel_file(file: Path, sheet_name: int | str = 0) -> pd.DataFrame:
    suffix = file.suffix.lower()
    if suffix == ".xlsx":
        return pd.read_excel(file, sheet_name=sheet_name, engine="openpyxl")
    if suffix == ".xls":
        if importlib.util.find_spec("xlrd") is None:
            raise ImportError(
                "古いExcel形式（.xls）を読み込むには xlrd が必要です。"
                "xlsx形式で保存し直すか、requirements-optional.txt の内容をインストールしてください。"
            )
        return pd.read_excel(file, sheet_name=sheet_name, engine="xlrd")
    raise ValueError(f"Excelファイルではありません: {file.name}")


def read_sales_file(file: Path, *, excel_sheet_name: int | str = 0) -> pd.DataFrame:
    suffix = file.suffix.lower()
    if suffix == ".csv":
        return read_csv_with_fallback(file)
    if suffix in {".xlsx", ".xls"}:
        return read_excel_file(file, sheet_name=excel_sheet_name)
    raise ValueError(
        f"対応していない入力ファイル形式です: {file.name}。"
        "対応形式は .csv / .xlsx / .xls です。"
    )


def read_sales_files(input_path: Path, pattern: str = "*.csv", *, excel_sheet_name: int | str = 0) -> pd.DataFrame:
    input_files = find_sales_files(input_path, pattern)
    dataframes = []
    for file in input_files:
        logging.info("読み込み開始: %s", file.name)
        df = read_sales_file(file, excel_sheet_name=excel_sheet_name)
        df["source_file"] = file.name
        df["source_row"] = df.index + 2
        dataframes.append(df)
        logging.info("読み込み完了: %s (%s件)", file.name, len(df))
    merged_df = pd.concat(dataframes, ignore_index=True)
    merged_df.attrs["source_files"] = [file.name for file in input_files]
    return merged_df


def read_csv_files(input_dir: Path, pattern: str = "*.csv") -> pd.DataFrame:
    if not input_dir.exists():
        raise FileNotFoundError(f"入力フォルダが見つかりません: {input_dir}")

    csv_files = sorted(
        file
        for file in input_dir.glob(pattern)
        if file.is_file() and file.suffix.lower() == ".csv" and not file.name.startswith("sample_")
    )
    if not csv_files:
        raise FileNotFoundError(
            "読み込み対象のCSVファイルが見つかりません。"
            f"入力フォルダ: {input_dir}。"
            f"ファイル名パターン: {pattern}。"
            "sample_ で始まるCSVはサンプルとして除外されます。"
        )

    dataframes = []
    for file in csv_files:
        logging.info("読み込み開始: %s", file.name)
        df = read_csv_with_fallback(file)
        df["source_file"] = file.name
        df["source_row"] = df.index + 2
        dataframes.append(df)
        logging.info("読み込み完了: %s", file.name)
    return pd.concat(dataframes, ignore_index=True)


def normalize_columns(df: pd.DataFrame, column_aliases: dict | None = None) -> pd.DataFrame:
    normalized_df = df.copy()
    aliases = merge_column_aliases(column_aliases)
    rename_map = {}
    mapped_columns = {}
    issues = []

    for column in normalized_df.columns:
        normalized_column = aliases.get(str(column).strip())
        if normalized_column is None:
            continue
        if normalized_column in mapped_columns:
            issues.append(
                ValidationIssue(
                    "duplicate_column",
                    f"同じ意味の列が重複しています。どちらか一方にしてください: {mapped_columns[normalized_column]}, {column}",
                    "-",
                    None,
                )
            )
        mapped_columns[normalized_column] = column
        rename_map[column] = normalized_column

    if issues:
        raise DataValidationError(issues)
    return normalized_df.rename(columns=rename_map)


def make_issue(row: pd.Series, issue: str, message: str, fallback_index: int) -> ValidationIssue:
    source_file = row.get("source_file")
    source_row = row.get("source_row")
    if pd.isna(source_file):
        source_file = "-"
    source_row_value = fallback_index + 2 if pd.isna(source_row) else int(source_row)
    return ValidationIssue(issue, message, str(source_file), source_row_value)


def validate_data(df: pd.DataFrame, column_aliases: dict | None = None) -> pd.DataFrame:
    df = normalize_columns(df, column_aliases)
    missing_columns = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing_columns:
        raise DataValidationError(
            [
                ValidationIssue(
                    "missing_column",
                    f"必須列が不足しています。不足列: {', '.join(missing_columns)}。必要な列: {', '.join(REQUIRED_COLUMNS)}",
                    "-",
                    None,
                )
            ]
        )

    validated_df = df.copy()
    date_text = validated_df["date"].astype("string").str.strip()
    invalid_date_format = ~date_text.str.fullmatch(r"\d{4}-\d{2}-\d{2}", na=False)
    validated_df["date"] = pd.to_datetime(date_text, format="%Y-%m-%d", errors="coerce")
    validated_df["quantity"] = pd.to_numeric(validated_df["quantity"], errors="coerce")
    validated_df["unit_price"] = pd.to_numeric(validated_df["unit_price"], errors="coerce")

    issues = []
    for index, row in validated_df[invalid_date_format | validated_df["date"].isna()].iterrows():
        issues.append(make_issue(row, "invalid_date", "日付に不正な値があります。YYYY-MM-DD 形式で入力してください。", index))
    for index, row in validated_df[validated_df["quantity"].isna()].iterrows():
        issues.append(make_issue(row, "invalid_quantity", "数量に不正な値があります。0以上の数値で入力してください。", index))
    for index, row in validated_df[validated_df["unit_price"].isna()].iterrows():
        issues.append(make_issue(row, "invalid_unit_price", "単価に不正な値があります。0以上の数値で入力してください。", index))
    for index, row in validated_df[validated_df["quantity"] < 0].iterrows():
        issues.append(make_issue(row, "negative_quantity", "数量にマイナスの値があります。0以上の数値で入力してください。", index))
    for index, row in validated_df[validated_df["unit_price"] < 0].iterrows():
        issues.append(make_issue(row, "negative_unit_price", "単価にマイナスの値があります。0以上の数値で入力してください。", index))

    if issues:
        raise DataValidationError(issues)

    validated_df["amount"] = validated_df["quantity"] * validated_df["unit_price"]
    return validated_df


def validate_date_option(value: str | None, option_name: str) -> pd.Timestamp | None:
    if value is None:
        return None
    try:
        return pd.to_datetime(value, format="%Y-%m-%d")
    except ValueError as exc:
        raise ValueError(f"{option_name} は YYYY-MM-DD 形式で指定してください。例: 2026-04-01") from exc


def filter_data(
    df: pd.DataFrame,
    *,
    month: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    product: str | None = None,
    category: str | None = None,
) -> pd.DataFrame:
    if month and (start_date or end_date):
        raise ValueError("--month と --start-date/--end-date は同時に指定できません。")

    filtered_df = df.copy()
    if month:
        filtered_df = filtered_df[filtered_df["date"].dt.strftime("%Y-%m") == month].copy()

    start = validate_date_option(start_date, "--start-date")
    end = validate_date_option(end_date, "--end-date")
    if start is not None and end is not None and start > end:
        raise ValueError("開始日が終了日より後になっています。開始日と終了日を確認してください。")
    if start is not None:
        filtered_df = filtered_df[filtered_df["date"] >= start].copy()
    if end is not None:
        filtered_df = filtered_df[filtered_df["date"] <= end].copy()
    if product:
        filtered_df = filtered_df[filtered_df["product"] == product].copy()
    if category:
        if "category" not in filtered_df.columns:
            raise ValueError("カテゴリで絞り込むには category 列が必要です。")
        filtered_df = filtered_df[filtered_df["category"] == category].copy()

    if filtered_df.empty:
        raise ValueError("指定条件に一致するデータがありません。日付範囲、商品、カテゴリの指定を確認してください。")
    return filtered_df


def create_summary(df: pd.DataFrame, group_by: str = "product") -> pd.DataFrame:
    if group_by not in GROUP_BY_COLUMNS:
        raise ValueError(f"未対応の集計単位です: {group_by}")
    if group_by not in df.columns:
        raise ValueError(f"{GROUP_BY_COLUMNS[group_by]}集計に必要な列がありません: {group_by}")

    summary_df = (
        df.groupby(group_by, as_index=False)
        .agg(
            total_amount=("amount", "sum"),
            transaction_count=("amount", "size"),
            total_quantity=("quantity", "sum"),
        )
        .sort_values(["total_amount", group_by], ascending=[False, True])
        .reset_index(drop=True)
    )
    summary_df["rank"] = range(1, len(summary_df) + 1)
    summary_df["average_unit_price"] = summary_df.apply(
        lambda row: row["total_amount"] / row["total_quantity"] if row["total_quantity"] else 0,
        axis=1,
    )
    total_amount = summary_df["total_amount"].sum()
    summary_df["total_amount_ratio"] = summary_df["total_amount"] / total_amount if total_amount else 0
    return summary_df[["rank", group_by, "total_amount", "transaction_count", "total_quantity", "average_unit_price", "total_amount_ratio"]]


def create_summaries(df: pd.DataFrame, group_by: str, all_summaries: bool = False) -> dict[str, pd.DataFrame]:
    group_keys = list(GROUP_BY_COLUMNS) if all_summaries else [group_by]
    return {key: create_summary(df, key) for key in group_keys}


def create_monthly_trend(df: pd.DataFrame) -> pd.DataFrame:
    trend_df = df.copy()
    trend_df["month"] = trend_df["date"].dt.strftime("%Y-%m")
    return (
        trend_df.groupby("month", as_index=False)
        .agg(total_quantity=("quantity", "sum"), total_amount=("amount", "sum"))
        .sort_values("month")
    )


def create_daily_trend(df: pd.DataFrame) -> pd.DataFrame:
    trend_df = df.copy()
    trend_df["date"] = trend_df["date"].dt.strftime("%Y-%m-%d")
    daily_df = (
        trend_df.groupby("date", as_index=False)
        .agg(
            total_amount=("amount", "sum"),
            transaction_count=("amount", "size"),
            total_quantity=("quantity", "sum"),
        )
        .sort_values("date")
        .reset_index(drop=True)
    )
    daily_df["average_unit_price"] = daily_df.apply(
        lambda row: row["total_amount"] / row["total_quantity"] if row["total_quantity"] else 0,
        axis=1,
    )
    return daily_df[["date", "total_amount", "transaction_count", "total_quantity", "average_unit_price"]]


def create_uncategorized_rows(df: pd.DataFrame) -> pd.DataFrame:
    output_columns = ["source_file", "source_row", "date", "product", "category", "quantity", "unit_price", "amount", "reason"]
    if df.empty:
        return pd.DataFrame(columns=output_columns)

    rows = []
    for _, row in df.iterrows():
        reasons: list[str] = []
        product_value = "" if pd.isna(row.get("product")) else str(row.get("product")).strip()
        category_value = "" if pd.isna(row.get("category")) else str(row.get("category")).strip()
        if not product_value:
            reasons.append("商品名が空欄")
        if not category_value:
            reasons.append("カテゴリが空欄")
        if product_value == "未設定":
            reasons.append("商品名が未設定")
        if category_value == "未分類":
            reasons.append("カテゴリが未分類")
        if category_value == "その他":
            reasons.append("カテゴリがその他")
        if not reasons:
            continue

        date_value = row.get("date")
        rows.append(
            {
                "source_file": row.get("source_file", ""),
                "source_row": row.get("source_row", ""),
                "date": date_value.strftime("%Y-%m-%d") if hasattr(date_value, "strftime") else date_value,
                "product": row.get("product", ""),
                "category": row.get("category", ""),
                "quantity": row.get("quantity", ""),
                "unit_price": row.get("unit_price", ""),
                "amount": row.get("amount", ""),
                "reason": " / ".join(reasons),
            }
        )
    return pd.DataFrame(rows, columns=output_columns)


def _month_period_label(df: pd.DataFrame, month: str | None = None) -> tuple[str, str]:
    if month:
        period = pd.Period(month, freq="M")
        return month, f"{period.start_time.strftime('%Y-%m-%d')} ～ {period.end_time.strftime('%Y-%m-%d')}"
    if df.empty:
        return "", ""
    start = df["date"].min().strftime("%Y-%m-%d")
    end = df["date"].max().strftime("%Y-%m-%d")
    month_label = df["date"].min().strftime("%Y-%m")
    return month_label, f"{start} ～ {end}"


def summarize_month_metrics(
    df: pd.DataFrame,
    month: str | None = None,
    *,
    uncategorized_count: int = 0,
    error_count: int = 0,
) -> dict[str, object]:
    month_label, period_label = _month_period_label(df, month)
    total_amount = float(df["amount"].sum()) if not df.empty and "amount" in df.columns else 0
    detail_count = int(len(df))
    total_quantity = float(df["quantity"].sum()) if not df.empty and "quantity" in df.columns else 0
    average_unit_price = total_amount / total_quantity if total_quantity else 0
    target_days = int(df["date"].dt.normalize().nunique()) if not df.empty and "date" in df.columns else 0
    product_count = int(df["product"].nunique()) if not df.empty and "product" in df.columns else 0
    category_count = int(df["category"].nunique()) if not df.empty and "category" in df.columns else 0

    top_product = ""
    top_product_amount = 0
    if not df.empty and "product" in df.columns:
        product_summary = create_summary(df, "product")
        if not product_summary.empty:
            top_product = product_summary.iloc[0]["product"]
            top_product_amount = float(product_summary.iloc[0]["total_amount"])

    top_category = ""
    top_category_amount = 0
    if not df.empty and "category" in df.columns:
        category_summary = create_summary(df, "category")
        if not category_summary.empty:
            top_category = category_summary.iloc[0]["category"]
            top_category_amount = float(category_summary.iloc[0]["total_amount"])

    return {
        "target_month": month_label,
        "period": period_label,
        "total_amount": total_amount,
        "detail_count": detail_count,
        "total_quantity": total_quantity,
        "average_unit_price": average_unit_price,
        "target_days": target_days,
        "product_count": product_count,
        "category_count": category_count,
        "top_product": top_product,
        "top_product_amount": top_product_amount,
        "top_category": top_category,
        "top_category_amount": top_category_amount,
        "uncategorized_count": uncategorized_count,
        "error_count": error_count,
        "review_count": uncategorized_count + error_count,
    }


def create_month_end_summary(
    df: pd.DataFrame,
    month: str | None = None,
    *,
    uncategorized_count: int = 0,
    error_count: int = 0,
) -> pd.DataFrame:
    metrics = summarize_month_metrics(df, month, uncategorized_count=uncategorized_count, error_count=error_count)
    return pd.DataFrame(
        [
            ("対象月", metrics["target_month"]),
            ("集計期間", metrics["period"]),
            ("売上合計", metrics["total_amount"]),
            ("明細件数", metrics["detail_count"]),
            ("数量合計", metrics["total_quantity"]),
            ("平均単価", metrics["average_unit_price"]),
            ("対象日数", metrics["target_days"]),
            ("商品数", metrics["product_count"]),
            ("カテゴリ数", metrics["category_count"]),
            ("売上トップ商品", metrics["top_product"]),
            ("売上トップ商品の売上", metrics["top_product_amount"]),
            ("売上トップカテゴリ", metrics["top_category"]),
            ("売上トップカテゴリの売上", metrics["top_category_amount"]),
            ("未分類データ件数", metrics["uncategorized_count"]),
            ("エラー行件数", metrics["error_count"]),
            ("確認が必要な件数", metrics["review_count"]),
        ],
        columns=["項目", "内容"],
    )


def create_previous_month_comparison(
    current_df: pd.DataFrame,
    previous_df: pd.DataFrame,
    month: str | None = None,
) -> pd.DataFrame:
    current = summarize_month_metrics(current_df, month)
    previous = summarize_month_metrics(previous_df)
    metric_pairs = [
        ("売上合計", "total_amount"),
        ("明細件数", "detail_count"),
        ("数量合計", "total_quantity"),
        ("平均単価", "average_unit_price"),
        ("対象日数", "target_days"),
        ("商品数", "product_count"),
        ("カテゴリ数", "category_count"),
    ]
    rows = []
    for label, key in metric_pairs:
        current_value = current[key]
        previous_value = previous[key]
        difference = current_value - previous_value
        change_rate = "比較不可" if not previous_value else difference / previous_value
        rows.append(
            {
                "指標": label,
                "当月": current_value,
                "前月": previous_value,
                "差分": difference,
                "増減率": change_rate,
            }
        )
    if previous["detail_count"] == 0:
        rows.append({"指標": "備考", "当月": "", "前月": "前月データなし", "差分": "", "増減率": "比較不可"})
    return pd.DataFrame(rows, columns=["指標", "当月", "前月", "差分", "増減率"])


def write_summary_csvs(summary_dfs: dict[str, pd.DataFrame], output_dir: Path, prefix: str = "summary") -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    output_files = []
    for group_by, summary_df in summary_dfs.items():
        output_file = output_dir / f"{prefix}_{group_by}.csv"
        summary_df.rename(columns=SUMMARY_COLUMN_LABELS).to_csv(output_file, index=False, encoding="utf-8-sig")
        output_files.append(output_file)
    return output_files


def format_report_period(detail_df: pd.DataFrame, month: str | None, start_date: str | None = None, end_date: str | None = None) -> str:
    if month:
        return f"対象月: {month}"
    if start_date or end_date:
        return f"対象期間: {start_date or '開始指定なし'} - {end_date or '終了指定なし'}"
    start = detail_df["date"].min().strftime("%Y-%m-%d")
    end = detail_df["date"].max().strftime("%Y-%m-%d")
    return f"対象期間: {start} - {end}"


def add_total_row(df: pd.DataFrame, label_column: str, total_columns: list[str]) -> pd.DataFrame:
    total_row = {column: "" for column in df.columns}
    total_row[label_column] = "合計"
    for column in total_columns:
        if column in df.columns:
            total_row[column] = df[column].sum()
    if "売上構成比" in df.columns:
        total_row["売上構成比"] = 1 if total_row.get("金額合計", total_row.get("売上合計", 0)) else 0
    return pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)


def prepare_detail_for_excel(detail_df: pd.DataFrame) -> pd.DataFrame:
    detail_columns = [column for column in DETAIL_COLUMN_LABELS if column in detail_df.columns]
    excel_df = detail_df.loc[:, detail_columns].copy()
    excel_df["date"] = excel_df["date"].dt.strftime("%Y-%m-%d")
    excel_df = excel_df.rename(columns=DETAIL_COLUMN_LABELS)
    return add_total_row(excel_df, excel_df.columns[0], ["数量", "金額"])


def prepare_summary_for_excel(summary_df: pd.DataFrame, group_by: str) -> pd.DataFrame:
    excel_df = summary_df.rename(columns=SUMMARY_COLUMN_LABELS).copy()
    excel_df = add_total_row(excel_df, SUMMARY_COLUMN_LABELS[group_by], ["売上合計", "件数", "数量合計"])
    total_amount = excel_df.iloc[-1].get("売上合計", 0)
    total_quantity = excel_df.iloc[-1].get("数量合計", 0)
    if "平均単価" in excel_df.columns:
        excel_df.at[excel_df.index[-1], "平均単価"] = total_amount / total_quantity if total_quantity else 0
    if "売上構成比" in excel_df.columns:
        excel_df.at[excel_df.index[-1], "売上構成比"] = 1 if total_amount else 0
    return excel_df


def prepare_monthly_trend_for_excel(trend_df: pd.DataFrame) -> pd.DataFrame:
    excel_df = trend_df.rename(columns=MONTHLY_TREND_COLUMN_LABELS).copy()
    return add_total_row(excel_df, "月", ["数量合計", "金額合計"])


def prepare_daily_trend_for_excel(trend_df: pd.DataFrame) -> pd.DataFrame:
    excel_df = trend_df.rename(columns=DAILY_TREND_COLUMN_LABELS).copy()
    excel_df = add_total_row(excel_df, "日付", ["売上合計", "件数", "数量合計"])
    total_amount = excel_df.iloc[-1].get("売上合計", 0)
    total_quantity = excel_df.iloc[-1].get("数量合計", 0)
    if "平均単価" in excel_df.columns:
        excel_df.at[excel_df.index[-1], "平均単価"] = total_amount / total_quantity if total_quantity else 0
    return excel_df


def prepare_month_end_summary_for_excel(summary_df: pd.DataFrame) -> pd.DataFrame:
    return summary_df.copy()


def prepare_previous_month_comparison_for_excel(comparison_df: pd.DataFrame) -> pd.DataFrame:
    return comparison_df.copy()


def prepare_uncategorized_for_excel(uncategorized_df: pd.DataFrame) -> pd.DataFrame:
    if uncategorized_df.empty:
        return pd.DataFrame([{"メッセージ": "確認が必要な未分類データはありません"}])
    return uncategorized_df.rename(
        columns={
            "source_file": "元ファイル名",
            "source_row": "行番号",
            "date": "date",
            "product": "product",
            "category": "category",
            "quantity": "quantity",
            "unit_price": "unit_price",
            "amount": "amount",
            "reason": "確認理由",
        }
    )


def prepare_validation_errors_for_excel(error_df: pd.DataFrame) -> pd.DataFrame:
    if error_df.empty:
        return pd.DataFrame([{"メッセージ": "検証エラーはありません"}])
    return error_df.rename(
        columns={
            "source_file": "元ファイル名",
            "source_row": "行番号",
            "message": "エラー内容",
            "date": "date",
            "product": "product",
            "category": "category",
            "quantity": "quantity",
            "unit_price": "unit_price",
            "amount": "amount",
        }
    )


def build_run_conditions(
    *,
    month: str | None,
    start_date: str | None,
    end_date: str | None,
    product: str | None,
    category: str | None,
    group_by: str,
    input_dir: Path,
    output_dir: Path,
    pattern: str,
    all_summaries: bool,
    monthly_trend: bool,
    output_name: str | None,
    summary_csv_dir: Path | None,
    detail_count: int,
) -> pd.DataFrame:
    return pd.DataFrame(
        [
            ("実行日時", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("対象月", month or ""),
            ("開始日", start_date or ""),
            ("終了日", end_date or ""),
            ("商品絞り込み", product or ""),
            ("カテゴリ絞り込み", category or ""),
            ("集計単位", "商品別・カテゴリ別" if all_summaries else GROUP_BY_COLUMNS[group_by]),
            ("月別推移", "出力" if monthly_trend else "未出力"),
            ("入力フォルダ", str(input_dir)),
            ("出力フォルダ", str(output_dir)),
            ("ファイル名パターン", pattern),
            ("出力ファイル名", output_name or "自動"),
            ("集計CSV出力", str(summary_csv_dir) if summary_csv_dir else "未出力"),
            ("明細件数", detail_count),
        ],
        columns=["項目", "値"],
    )


def format_excel_sheet(
    worksheet,
    title: str,
    period_label: str,
    number_columns: set[str],
    percent_columns: set[str] | None = None,
    style_config: dict | None = None,
) -> None:
    style = merge_style_config(style_config)
    percent_columns = percent_columns or set()
    worksheet["A1"] = title
    worksheet["A1"].font = Font(bold=True, size=style["title_size"])
    worksheet["A2"] = period_label
    worksheet.freeze_panes = "A5"
    worksheet.sheet_view.showGridLines = False

    header_row = 4
    worksheet.auto_filter.ref = f"A{header_row}:{worksheet.cell(header_row, worksheet.max_column).coordinate}"
    for cell in worksheet[header_row]:
        cell.font = HEADER_FONT
        cell.fill = PatternFill(fill_type="solid", fgColor=style["header_fill"])
        cell.alignment = Alignment(horizontal="center")

    header_by_column = {cell.column_letter: cell.value for cell in worksheet[header_row] if cell.value is not None}
    for column_letter, header in header_by_column.items():
        if header in number_columns:
            for cell in worksheet[column_letter][header_row:]:
                cell.number_format = NUMBER_FORMAT
        if header in percent_columns:
            for cell in worksheet[column_letter][header_row:]:
                cell.number_format = PERCENT_FORMAT

    if any(cell.value == "合計" for cell in worksheet[worksheet.max_row]):
        for cell in worksheet[worksheet.max_row]:
            cell.font = TOTAL_FONT
            cell.fill = PatternFill(fill_type="solid", fgColor=style["total_fill"])

    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 50)


def format_conditions_sheet(worksheet, style_config: dict | None = None) -> None:
    style = merge_style_config(style_config)
    worksheet["A1"] = "実行条件"
    worksheet["A1"].font = Font(bold=True, size=style["title_size"])
    worksheet.sheet_view.showGridLines = False
    for cell in worksheet[3]:
        cell.font = HEADER_FONT
        cell.fill = PatternFill(fill_type="solid", fgColor=style["header_fill"])
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 60)


def format_key_value_sheet(worksheet, title: str, style_config: dict | None = None) -> None:
    style = merge_style_config(style_config)
    worksheet["A1"] = title
    worksheet["A1"].font = Font(bold=True, size=style["title_size"])
    worksheet.sheet_view.showGridLines = False
    for cell in worksheet[3]:
        cell.font = HEADER_FONT
        cell.fill = PatternFill(fill_type="solid", fgColor=style["header_fill"])
        cell.alignment = Alignment(horizontal="center")
    for row in worksheet.iter_rows(min_row=4):
        if row and str(row[0].value or "").endswith(("売上", "合計", "単価")):
            row[1].number_format = NUMBER_FORMAT
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 14), 60)


def format_comparison_sheet(worksheet, style_config: dict | None = None) -> None:
    style = merge_style_config(style_config)
    worksheet["A1"] = "前月比較"
    worksheet["A1"].font = Font(bold=True, size=style["title_size"])
    worksheet.sheet_view.showGridLines = False
    for cell in worksheet[3]:
        cell.font = HEADER_FONT
        cell.fill = PatternFill(fill_type="solid", fgColor=style["header_fill"])
        cell.alignment = Alignment(horizontal="center")
    headers = {cell.value: cell.column_letter for cell in worksheet[3] if cell.value}
    for header in ("当月", "前月", "差分"):
        column_letter = headers.get(header)
        if column_letter:
            for cell in worksheet[column_letter][3:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = NUMBER_FORMAT
    rate_column = headers.get("増減率")
    if rate_column:
        for cell in worksheet[rate_column][3:]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = PERCENT_FORMAT
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 50)


def format_plain_table_sheet(worksheet, title: str, style_config: dict | None = None) -> None:
    style = merge_style_config(style_config)
    worksheet["A1"] = title
    worksheet["A1"].font = Font(bold=True, size=style["title_size"])
    worksheet.sheet_view.showGridLines = False
    for cell in worksheet[3]:
        cell.font = HEADER_FONT
        cell.fill = PatternFill(fill_type="solid", fgColor=style["header_fill"])
        cell.alignment = Alignment(horizontal="center")
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 60)


def add_summary_bar_chart(worksheet, title: str, style_config: dict | None = None) -> None:
    if worksheet.max_row <= 5:
        return
    style = merge_style_config(style_config)
    chart = BarChart()
    chart.title = title
    chart.y_axis.title = "金額"
    chart.x_axis.title = worksheet["B4"].value
    data = Reference(worksheet, min_col=3, min_row=4, max_row=worksheet.max_row - 1)
    categories = Reference(worksheet, min_col=2, min_row=5, max_row=worksheet.max_row - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = style["chart_height"]
    chart.width = style["chart_width"]
    worksheet.add_chart(chart, "F4")


def add_monthly_line_chart(worksheet, style_config: dict | None = None) -> None:
    if worksheet.max_row <= 5:
        return
    style = merge_style_config(style_config)
    chart = LineChart()
    chart.title = "月別売上推移"
    chart.y_axis.title = "金額"
    chart.x_axis.title = "月"
    data = Reference(worksheet, min_col=3, min_row=4, max_row=worksheet.max_row - 1)
    categories = Reference(worksheet, min_col=1, min_row=5, max_row=worksheet.max_row - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = style["chart_height"]
    chart.width = style["chart_width"]
    worksheet.add_chart(chart, "E4")


def add_daily_line_chart(worksheet, style_config: dict | None = None) -> None:
    if worksheet.max_row <= 5:
        return
    style = merge_style_config(style_config)
    chart = LineChart()
    chart.title = "日別売上推移"
    chart.y_axis.title = "売上合計"
    chart.x_axis.title = "日付"
    data = Reference(worksheet, min_col=2, min_row=4, max_row=worksheet.max_row - 1)
    categories = Reference(worksheet, min_col=1, min_row=5, max_row=worksheet.max_row - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = style["chart_height"]
    chart.width = style["chart_width"]
    worksheet.add_chart(chart, "G4")


def resolve_output_file(output_dir: Path, month: str | None, output_name: str | None) -> Path:
    if output_name:
        output_path = Path(output_name)
        if output_path.suffix.lower() != ".xlsx":
            output_path = output_path.with_suffix(".xlsx")
        if not output_path.is_absolute():
            output_path = output_dir / output_path
        return output_path
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    period_label = month.replace("-", "") if month else "custom"
    return output_dir / f"monthly_report_{period_label}_{timestamp}.xlsx"


def save_to_excel(
    detail_df: pd.DataFrame,
    summary_dfs: dict[str, pd.DataFrame],
    output_dir: Path,
    month: str | None,
    group_by: str = "product",
    *,
    input_dir: Path | None = None,
    pattern: str = "*.csv",
    all_summaries: bool = False,
    start_date: str | None = None,
    end_date: str | None = None,
    product: str | None = None,
    category: str | None = None,
    daily_trend_df: pd.DataFrame | None = None,
    monthly_trend_df: pd.DataFrame | None = None,
    month_end_summary_df: pd.DataFrame | None = None,
    previous_month_comparison_df: pd.DataFrame | None = None,
    uncategorized_df: pd.DataFrame | None = None,
    validation_error_df: pd.DataFrame | None = None,
    output_name: str | None = None,
    charts: bool = False,
    summary_csv_dir: Path | None = None,
    style_config: dict | None = None,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    output_file = resolve_output_file(output_dir, month, output_name)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    period_label = format_report_period(detail_df, month, start_date, end_date)
    conditions_df = build_run_conditions(
        month=month,
        start_date=start_date,
        end_date=end_date,
        product=product,
        category=category,
        group_by=group_by,
        input_dir=input_dir or Path(""),
        output_dir=output_dir,
        pattern=pattern,
        all_summaries=all_summaries,
        monthly_trend=monthly_trend_df is not None,
        output_name=output_name,
        summary_csv_dir=summary_csv_dir,
        detail_count=len(detail_df),
    )

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        if month_end_summary_df is not None:
            prepare_month_end_summary_for_excel(month_end_summary_df).to_excel(writer, sheet_name="月末サマリー", index=False, startrow=2)
        if previous_month_comparison_df is not None:
            prepare_previous_month_comparison_for_excel(previous_month_comparison_df).to_excel(writer, sheet_name="前月比較", index=False, startrow=2)
        prepare_detail_for_excel(detail_df).to_excel(writer, sheet_name="詳細データ", index=False, startrow=3)
        for summary_group, summary_df in summary_dfs.items():
            sheet_name = "集計結果" if len(summary_dfs) == 1 else f"{GROUP_BY_COLUMNS[summary_group]}集計"
            prepare_summary_for_excel(summary_df, summary_group).to_excel(writer, sheet_name=sheet_name, index=False, startrow=3)
        if daily_trend_df is not None:
            prepare_daily_trend_for_excel(daily_trend_df).to_excel(writer, sheet_name="日別推移", index=False, startrow=3)
        if monthly_trend_df is not None:
            prepare_monthly_trend_for_excel(monthly_trend_df).to_excel(writer, sheet_name="月別推移", index=False, startrow=3)
        prepare_uncategorized_for_excel(uncategorized_df if uncategorized_df is not None else pd.DataFrame()).to_excel(
            writer, sheet_name="未分類データ", index=False, startrow=2
        )
        prepare_validation_errors_for_excel(validation_error_df if validation_error_df is not None else pd.DataFrame()).to_excel(
            writer, sheet_name="エラー行一覧", index=False, startrow=2
        )
        conditions_df.to_excel(writer, sheet_name="実行条件", index=False, startrow=2)

        if month_end_summary_df is not None:
            format_key_value_sheet(writer.sheets["月末サマリー"], "月末サマリー", style_config)
        if previous_month_comparison_df is not None:
            format_comparison_sheet(writer.sheets["前月比較"], style_config)
        format_excel_sheet(writer.sheets["詳細データ"], "月次売上レポート - 詳細データ", period_label, {"数量", "単価", "元CSV行", "金額"}, style_config=style_config)
        for summary_group in summary_dfs:
            sheet_name = "集計結果" if len(summary_dfs) == 1 else f"{GROUP_BY_COLUMNS[summary_group]}集計"
            format_excel_sheet(
                writer.sheets[sheet_name],
                f"月次売上レポート - {GROUP_BY_COLUMNS[summary_group]}集計結果",
                period_label,
                {"順位", "売上合計", "件数", "数量合計", "平均単価"},
                {"売上構成比"},
                style_config=style_config,
            )
            if charts:
                add_summary_bar_chart(writer.sheets[sheet_name], f"{GROUP_BY_COLUMNS[summary_group]}売上", style_config)
        if daily_trend_df is not None:
            format_excel_sheet(
                writer.sheets["日別推移"],
                "月次売上レポート - 日別推移",
                period_label,
                {"売上合計", "件数", "数量合計", "平均単価"},
                style_config=style_config,
            )
            if charts:
                add_daily_line_chart(writer.sheets["日別推移"], style_config)
        if monthly_trend_df is not None:
            format_excel_sheet(writer.sheets["月別推移"], "月次売上レポート - 月別推移", period_label, {"数量合計", "金額合計"}, style_config=style_config)
            if charts:
                add_monthly_line_chart(writer.sheets["月別推移"], style_config)
        format_plain_table_sheet(writer.sheets["未分類データ"], "未分類データ", style_config)
        format_plain_table_sheet(writer.sheets["エラー行一覧"], "エラー行一覧", style_config)
        format_conditions_sheet(writer.sheets["実行条件"], style_config)

    return output_file


def cleanup_old_reports(output_dir: Path, keep_reports: int | None) -> list[Path]:
    if keep_reports is None:
        return []
    report_files = sorted(output_dir.glob("monthly_report_*.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True)
    delete_targets = report_files[keep_reports:]
    for file in delete_targets:
        file.unlink()
    return delete_targets
