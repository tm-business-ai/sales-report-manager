import argparse
import csv
import json
import logging
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

import main


def test_validate_keep_reports_accepts_positive_integer() -> None:
    assert main.validate_keep_reports("3") == 3


@pytest.mark.parametrize("value", ["0", "-1", "abc"])
def test_validate_keep_reports_rejects_invalid_values(value: str) -> None:
    with pytest.raises(argparse.ArgumentTypeError):
        main.validate_keep_reports(value)


def test_validate_month_accepts_yyyy_mm() -> None:
    assert main.validate_month("2026-04") == "2026-04"
    assert main.validate_month(None) is None


def test_validate_options_rejects_month_and_date_range_together() -> None:
    with pytest.raises(ValueError, match="同時に指定できません"):
        main.validate_options("2026-04", "2026-04-01", None, "product")


def test_validate_options_rejects_start_date_after_end_date() -> None:
    with pytest.raises(ValueError, match="開始日が終了日より後"):
        main.validate_options(None, "2026-04-30", "2026-04-01", "product")


def test_validate_options_start_date_after_end_date_has_fix_example() -> None:
    with pytest.raises(ValueError) as exc_info:
        main.validate_options(None, "2026-04-30", "2026-04-01", "product")

    message = str(exc_info.value)
    assert "修正例" in message
    assert "2026-04-01" in message


def test_load_config_merges_preset_and_validates_unknown_key(tmp_path: Path) -> None:
    config_file = tmp_path / "config.json"
    config_file.write_text(
        json.dumps(
            {
                "input_dir": "base_input",
                "pattern": "*.csv",
                "presets": {
                    "april": {
                        "month": "2026-04",
                        "pattern": "sales_*.csv",
                    }
                },
            }
        ),
        encoding="utf-8",
    )

    config = main.load_config(config_file, "april")

    assert config["input_dir"] == "base_input"
    assert config["month"] == "2026-04"
    assert config["pattern"] == "sales_*.csv"

    bad_config = tmp_path / "bad_config.json"
    bad_config.write_text('{"unknown": true}', encoding="utf-8")
    with pytest.raises(ValueError, match="未対応"):
        main.load_config(bad_config, None)


def test_list_presets_returns_sorted_names(tmp_path: Path) -> None:
    config_file = tmp_path / "config.json"
    config_file.write_text(
        json.dumps({"presets": {"z": {"month": "2026-05"}, "a": {"month": "2026-04"}}}),
        encoding="utf-8",
    )

    assert main.list_presets(config_file) == ("a", "z")


def test_setup_logging_uses_rotation(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(main, "LOG_DIR", tmp_path / "logs")
    monkeypatch.setattr(main, "LOG_FILE", tmp_path / "logs" / "app.log")

    main.setup_logging(max_bytes=100, backup_count=1)
    logger = logging.getLogger()

    assert logger.handlers
    assert getattr(logger.handlers[0], "maxBytes") == 100


def test_help_does_not_create_log_file(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    log_file = tmp_path / "logs" / "app.log"
    monkeypatch.setattr(main, "LOG_DIR", tmp_path / "logs")
    monkeypatch.setattr(main, "LOG_FILE", log_file)
    monkeypatch.setattr(sys, "argv", ["main.py", "--help"])

    with pytest.raises(SystemExit) as exc_info:
        main.main()

    assert exc_info.value.code == 0
    assert not log_file.exists()


def test_check_environment_reports_missing_paths(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(main, "REQUIRED_MODULES", ())

    issues = main.check_environment(
        input_dir=tmp_path / "missing_input",
        output_dir=tmp_path / "missing_output",
        requirements_file=tmp_path / "missing_requirements.txt",
    )

    assert any("requirements.txt" in issue for issue in issues)
    assert any("missing_input" in issue for issue in issues)
    assert any("missing_output" in issue for issue in issues)


def test_main_check_setup_does_not_create_log_file(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    capsys: pytest.CaptureFixture[str],
) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    log_file = tmp_path / "logs" / "app.log"
    input_dir.mkdir()
    output_dir.mkdir()
    monkeypatch.setattr(main, "INPUT_DIR", input_dir)
    monkeypatch.setattr(main, "OUTPUT_DIR", output_dir)
    monkeypatch.setattr(main, "LOG_DIR", tmp_path / "logs")
    monkeypatch.setattr(main, "LOG_FILE", log_file)
    monkeypatch.setattr(main, "REQUIRED_MODULES", ())
    monkeypatch.setattr(sys, "argv", ["main.py", "--check-setup"])

    exit_code = main.main()
    captured = capsys.readouterr()

    assert exit_code == 0
    assert "OK" in captured.out
    assert not log_file.exists()


def test_write_input_template_creates_expected_columns(tmp_path: Path) -> None:
    output_file = main.write_input_template(tmp_path / "template.csv")

    text = output_file.read_text(encoding="utf-8-sig")
    assert text.splitlines()[0] == "date,product,category,quantity,unit_price"


def test_append_audit_log_writes_jsonl(tmp_path: Path) -> None:
    result = main.ReportRunResult(output_file=tmp_path / "report.xlsx", detail_count=2, summary_count=1)
    log_file = main.append_audit_log(
        status="success",
        options={"input_dir": tmp_path / "input", "output_dir": tmp_path / "output", "group_by": "product"},
        result=result,
        audit_log_file=tmp_path / "history.jsonl",
    )

    record = json.loads(log_file.read_text(encoding="utf-8").strip())
    assert record["status"] == "success"
    assert record["detail_count"] == 2
    assert record["output_file"].endswith("report.xlsx")
    assert record["warnings"] == []


def test_append_audit_log_writes_rerun_options_and_warnings(tmp_path: Path) -> None:
    result = main.ReportRunResult(
        output_file=None,
        detail_count=0,
        summary_count=0,
        dry_run=True,
        warnings=("出力対象の明細が0件です。条件または入力CSV / Excelを確認してください。",),
    )

    log_file = main.append_audit_log(
        status="success",
        options={
            "input_dir": tmp_path / "input",
            "output_dir": tmp_path / "output",
            "group_by": "category",
            "pattern": "sales*.csv",
            "all_summaries": True,
            "monthly_trend": True,
            "charts": False,
            "warning_amount_threshold": 5000,
            "column_aliases": {"sold_at": "date"},
            "style_config": {"header_fill": "FFFFFF"},
        },
        result=result,
        audit_log_file=tmp_path / "history.jsonl",
    )

    record = json.loads(log_file.read_text(encoding="utf-8").strip())
    assert record["pattern"] == "sales*.csv"
    assert record["all_summaries"] is True
    assert record["column_aliases"] == {"sold_at": "date"}
    assert record["warning_amount_threshold"] == 5000
    assert record["style"] == {"header_fill": "FFFFFF"}
    assert record["warnings"]


def test_read_audit_log_returns_recent_valid_json_records(tmp_path: Path) -> None:
    audit_file = tmp_path / "history.jsonl"
    audit_file.write_text(
        "\n".join(
            [
                json.dumps({"timestamp": "1", "status": "success"}),
                "not json",
                json.dumps({"timestamp": "2", "status": "error"}),
            ]
        ),
        encoding="utf-8",
    )

    records = main.read_audit_log(audit_file, limit=1)

    assert records == [{"timestamp": "2", "status": "error"}]


def test_normalize_legacy_text_repairs_mojibake() -> None:
    legacy_text = "險ｭ螳壹ヵ繧｡繧､繝ｫ"

    repaired = main.normalize_legacy_text(legacy_text)

    assert repaired == "設定ファイル"


def test_read_audit_log_normalizes_legacy_records(tmp_path: Path) -> None:
    audit_file = tmp_path / "history.jsonl"
    audit_file.write_text(
        json.dumps(
            {
                "timestamp": "2026-04-28T00:00:00",
                "status": "success",
                "warnings": ["繝ｭ繧ｰ縺ｯ縺ｾ縺縺ゅｊ縺ｾ縺帙ｓ"],
                "error": "險ｭ螳壹ヵ繧｡繧､繝ｫ",
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    records = main.read_audit_log(audit_file)

    assert records[0]["warnings"] == ["ログはまだありません"]
    assert records[0]["error"] == "設定ファイル"


def test_export_audit_log_csv_writes_flattened_rows(tmp_path: Path) -> None:
    output_file = main.export_audit_log_csv(
        [
            {
                "timestamp": "2026-04-28T00:00:00",
                "status": "success",
                "month": "2026-04",
                "detail_count": 10,
                "summary_count": 2,
                "warnings": ["警告1", "警告2"],
                "summary_csv_files": ["a.csv", "b.csv"],
                "column_aliases": {"sold_at": "date"},
                "style": {"header_fill": "FFFFFF"},
            }
        ],
        tmp_path / "audit.csv",
    )

    rows = list(csv.DictReader(output_file.open(encoding="utf-8-sig", newline="")))

    assert len(rows) == 1
    assert rows[0]["period"] == "2026-04"
    assert rows[0]["warning_count"] == "2"
    assert rows[0]["warnings"] == "警告1 | 警告2"
    assert rows[0]["summary_csv_files"] == "a.csv | b.csv"
    assert '"sold_at": "date"' in rows[0]["column_aliases"]


def test_filter_audit_records_supports_status_text_and_date_range() -> None:
    records = [
        {"timestamp": "2026-04-01T09:00:00", "status": "success", "output_file": "a.xlsx", "error": ""},
        {"timestamp": "2026-04-15T09:00:00", "status": "error", "output_file": "b.xlsx", "error": "boom"},
        {"timestamp": "2026-05-01T09:00:00", "status": "success", "output_file": "c.xlsx", "error": ""},
    ]

    filtered = main.filter_audit_records(
        records,
        filter_text="boom",
        filter_status="error",
        date_from="2026-04-01",
        date_to="2026-04-30",
    )

    assert filtered == [records[1]]


def test_export_audit_summary_csv_writes_single_summary_row(tmp_path: Path) -> None:
    output_file = main.export_audit_summary_csv(
        [
            {"timestamp": "2026-04-01T09:00:00", "status": "success", "warnings": ["a"], "dry_run": True},
            {"timestamp": "2026-04-02T09:00:00", "status": "validation_error", "warnings": []},
            {"timestamp": "2026-04-03T09:00:00", "status": "error", "warnings": ["b", "c"]},
        ],
        tmp_path / "audit_summary.csv",
    )

    rows = list(csv.DictReader(output_file.open(encoding="utf-8-sig", newline="")))

    assert len(rows) == 1
    assert rows[0]["total_runs"] == "3"
    assert rows[0]["success_count"] == "1"
    assert rows[0]["validation_error_count"] == "1"
    assert rows[0]["error_count"] == "1"
    assert rows[0]["warning_total"] == "3"
    assert rows[0]["dry_run_count"] == "1"


def test_export_audit_monthly_summary_csv_groups_by_month(tmp_path: Path) -> None:
    output_file = main.export_audit_monthly_summary_csv(
        [
            {"timestamp": "2026-04-01T09:00:00", "status": "success", "warnings": ["a"]},
            {"timestamp": "2026-04-02T09:00:00", "status": "error", "warnings": []},
            {"timestamp": "2026-05-01T09:00:00", "status": "validation_error", "warnings": ["b", "c"], "dry_run": True},
        ],
        tmp_path / "audit_monthly_summary.csv",
    )

    rows = list(csv.DictReader(output_file.open(encoding="utf-8-sig", newline="")))

    assert len(rows) == 2
    assert rows[0]["month"] == "2026-04"
    assert rows[0]["total_runs"] == "2"
    assert rows[1]["month"] == "2026-05"
    assert rows[1]["validation_error_count"] == "1"
    assert rows[1]["warning_total"] == "2"
    assert rows[1]["dry_run_count"] == "1"


def test_inspect_legacy_text_files_returns_preview_candidates(tmp_path: Path) -> None:
    gui_state_file = tmp_path / "gui_state.json"
    audit_log_file = tmp_path / "report_history.jsonl"
    alias_preset_file = tmp_path / "column_alias_presets.json"
    gui_state_file.write_text(json.dumps({"label": "險ｭ螳壹ヵ繧｡繧､繝ｫ"}, ensure_ascii=False), encoding="utf-8")
    audit_log_file.write_text(json.dumps({"error": "險ｭ螳壹ヵ繧｡繧､繝ｫ"}, ensure_ascii=False) + "\n", encoding="utf-8")
    alias_preset_file.write_text(json.dumps({"shop_a": {"note": "設定ファイル"}}, ensure_ascii=False), encoding="utf-8")

    changes = main.inspect_legacy_text_files(gui_state_file, audit_log_file, alias_preset_file)
    preview = main.format_legacy_text_preview(changes)

    assert changes[str(gui_state_file)]
    assert changes[str(audit_log_file)]
    assert changes[str(alias_preset_file)] == []
    assert "設定ファイル" in preview


def test_prune_audit_log_applies_keep_count_and_keep_days(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    audit_log_file = tmp_path / "report_history.jsonl"
    audit_log_file.write_text(
        "\n".join(
            [
                json.dumps({"timestamp": "2026-04-01T09:00:00", "status": "success"}),
                json.dumps({"timestamp": "2026-04-27T09:00:00", "status": "success"}),
                json.dumps({"timestamp": "2026-04-28T09:00:00", "status": "error"}),
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    class FrozenDateTime:
        @classmethod
        def now(cls):
            from datetime import datetime as real_datetime

            return real_datetime(2026, 4, 28, 12, 0, 0)

        @classmethod
        def fromisoformat(cls, value: str):
            from datetime import datetime as real_datetime

            return real_datetime.fromisoformat(value)

        @classmethod
        def strptime(cls, value: str, fmt: str):
            from datetime import datetime as real_datetime

            return real_datetime.strptime(value, fmt)

    monkeypatch.setattr(main, "datetime", FrozenDateTime)

    removed = main.prune_audit_log(audit_log_file, keep_count=1, keep_days=2)
    records = main.read_audit_log(audit_log_file, limit=10)

    assert removed == 2
    assert len(records) == 1
    assert records[0]["timestamp"] == "2026-04-28T09:00:00"


def test_backup_audit_log_creates_zip(tmp_path: Path) -> None:
    audit_log_file = tmp_path / "report_history.jsonl"
    audit_log_file.write_text('{"status":"success"}\n', encoding="utf-8")

    backup_file = main.backup_audit_log(audit_log_file, tmp_path / "backup")

    assert backup_file is not None
    assert backup_file.exists()
    assert backup_file.suffix == ".zip"


def test_detect_audit_anomalies_reports_failure_rate_and_missing_days(monkeypatch: pytest.MonkeyPatch) -> None:
    class FrozenDateTime:
        @classmethod
        def now(cls):
            from datetime import datetime as real_datetime

            return real_datetime(2026, 5, 5, 12, 0, 0)

        @classmethod
        def fromisoformat(cls, value: str):
            from datetime import datetime as real_datetime

            return real_datetime.fromisoformat(value)

    monkeypatch.setattr(main, "datetime", FrozenDateTime)

    anomaly = main.detect_audit_anomalies(
        [
            {"timestamp": "2026-05-01T09:00:00", "status": "error", "warnings": ["a"]},
            {"timestamp": "2026-05-02T09:00:00", "status": "validation_error", "warnings": ["b"]},
            {"timestamp": "2026-05-02T10:00:00", "status": "success", "warnings": ["c"]},
        ]
    )

    assert anomaly.failure_count == 2
    assert anomaly.consecutive_missing_days == 3
    assert anomaly.failure_rate > 0.3
    assert anomaly.alerts


def test_column_alias_presets_round_trip(tmp_path: Path) -> None:
    preset_file = tmp_path / "column_alias_presets.json"
    presets = {
        "shop_a": {
            "sold_at": "date",
            "item": "product",
            "qty": "quantity",
            "price": "unit_price",
        }
    }

    written_file = main.write_column_alias_presets(presets, preset_file)

    assert written_file == preset_file
    assert main.read_column_alias_presets(preset_file) == presets


def test_build_preview_returns_filtered_rows(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    input_dir.mkdir()
    (input_dir / "sales.csv").write_text(
        "date,product,category,quantity,unit_price\n"
        "2026-04-01,apple,fruit,10,120\n"
        "2026-05-02,orange,fruit,8,100\n",
        encoding="utf-8",
    )

    preview = main.build_preview(
        input_dir=input_dir,
        month="2026-04",
        group_by="product",
        pattern="sales*.csv",
        limit=5,
    )

    assert preview.total_count == 1
    assert "product" in preview.columns
    assert preview.rows[0][preview.columns.index("product")] == "apple"


def test_build_preview_accepts_configured_column_aliases(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    input_dir.mkdir()
    (input_dir / "sales.csv").write_text(
        "sold_at,item,dept,qty,price\n"
        "2026-04-01,apple,fruit,10,120\n",
        encoding="utf-8",
    )

    preview = main.build_preview(
        input_dir=input_dir,
        month="2026-04",
        group_by="product",
        pattern="sales*.csv",
        column_aliases={
            "sold_at": "date",
            "item": "product",
            "dept": "category",
            "qty": "quantity",
            "price": "unit_price",
        },
    )

    assert preview.total_count == 1
    assert "product" in preview.columns
    assert preview.rows[0][preview.columns.index("product")] == "apple"


def test_build_summary_preview_returns_summary_rows(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    input_dir.mkdir()
    (input_dir / "sales.csv").write_text(
        "date,product,category,quantity,unit_price\n"
        "2026-04-01,apple,fruit,10,120\n"
        "2026-04-02,apple,fruit,5,120\n"
        "2026-04-03,carrot,vegetable,3,90\n",
        encoding="utf-8",
    )

    preview = main.build_summary_preview(
        input_dir=input_dir,
        month="2026-04",
        group_by="product",
        pattern="sales*.csv",
        all_summaries=True,
    )

    assert preview.detail_count == 3
    assert set(preview.summaries) == {"product", "category", "daily"}
    assert preview.summaries["product"].total_count == 2
    assert preview.summaries["daily"].total_count == 3


def test_build_preview_no_matching_month_has_confirmation_points(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    input_dir.mkdir()
    (input_dir / "sales.csv").write_text(
        "date,product,category,quantity,unit_price\n2026-04-01,apple,fruit,10,120\n",
        encoding="utf-8",
    )

    with pytest.raises(ValueError) as exc_info:
        main.build_preview(input_dir=input_dir, month="2026-05", group_by="product", pattern="sales*.csv")

    message = str(exc_info.value)
    assert "対象月に一致する売上データがありません" in message
    assert "確認してください" in message
    assert "対象月: 2026-05" in message


def test_inspect_report_warnings_reports_empty_missing_category_and_high_amount() -> None:
    import pandas as pd

    empty_warnings = main.inspect_report_warnings(pd.DataFrame(columns=["category", "amount"]))
    data_warnings = main.inspect_report_warnings(
        pd.DataFrame(
            [
                {"category": "", "amount": 10},
                {"category": "fruit", "amount": 1_500_000},
            ]
        )
    )

    assert any("0件" in warning for warning in empty_warnings)
    assert any("カテゴリ" in warning for warning in data_warnings)
    assert any("1,000,000" in warning for warning in data_warnings)


def test_run_report_uses_configured_warning_amount_threshold(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    (input_dir / "sales.csv").write_text(
        "date,product,category,quantity,unit_price\n"
        "2026-04-01,apple,fruit,2,600\n",
        encoding="utf-8",
    )

    result = main.run_report(
        input_dir=input_dir,
        output_dir=output_dir,
        month="2026-04",
        group_by="product",
        keep_reports=None,
        pattern="sales*.csv",
        dry_run=True,
        warning_amount_threshold=1_000,
    )

    assert any("1,000" in warning for warning in result.warnings)


def test_main_preview_prints_rows(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    capsys: pytest.CaptureFixture[str],
) -> None:
    input_dir = tmp_path / "input"
    input_dir.mkdir()
    (input_dir / "sales.csv").write_text(
        "date,product,category,quantity,unit_price\n2026-04-01,apple,fruit,10,120\n",
        encoding="utf-8",
    )
    monkeypatch.setattr(sys, "argv", ["main.py", "--input-dir", str(input_dir), "--month", "2026-04", "--preview"])

    exit_code = main.main()
    captured = capsys.readouterr()

    assert exit_code == 0
    assert "apple" in captured.out
    assert "プレビュー件数" in captured.out


def test_run_report_creates_excel_charts_custom_name_and_summary_csv(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    csv_dir = tmp_path / "csv"
    input_dir.mkdir()
    (input_dir / "sales.csv").write_text(
        "date,product,category,quantity,unit_price\n"
        "2026-04-01,りんご,果物,10,120\n"
        "2026-05-02,みかん,果物,8,100\n",
        encoding="utf-8",
    )

    result = main.run_report(
        input_dir=input_dir,
        output_dir=output_dir,
        month=None,
        start_date="2026-04-01",
        end_date="2026-05-31",
        product=None,
        category="果物",
        group_by="product",
        keep_reports=None,
        pattern="sales*.csv",
        all_summaries=True,
        monthly_trend=True,
        charts=True,
        output_name="sales_report.xlsx",
        summary_csv_dir=csv_dir,
        summary_csv_prefix="sales",
    )

    assert result.output_file == output_dir / "sales_report.xlsx"
    workbook = load_workbook(result.output_file, data_only=True)
    assert workbook.sheetnames == ["月末サマリー", "前月比較", "詳細データ", "商品別集計", "カテゴリ別集計", "日別推移", "月別推移", "未分類データ", "エラー行一覧", "実行条件"]
    assert len(workbook["商品別集計"]._charts) == 1
    assert len(workbook["日別推移"]._charts) == 1
    assert len(workbook["月別推移"]._charts) == 1
    assert {path.name for path in result.summary_csv_files} == {"sales_product.csv", "sales_category.csv"}
    assert (csv_dir / "sales_product.csv").exists()
    assert result.detail_count == 2


def test_run_report_dry_run_does_not_create_excel_or_summary_csv(tmp_path: Path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    csv_dir = tmp_path / "csv"
    input_dir.mkdir()
    (input_dir / "sales.csv").write_text(
        "date,product,category,quantity,unit_price\n2026-04-01,りんご,果物,10,120\n",
        encoding="utf-8",
    )

    result = main.run_report(
        input_dir=input_dir,
        output_dir=output_dir,
        month="2026-04",
        group_by="product",
        keep_reports=None,
        dry_run=True,
        summary_csv_dir=csv_dir,
    )

    assert result.dry_run is True
    assert result.output_file is None
    assert result.summary_csv_files == ()
    assert not output_dir.exists()
    assert not csv_dir.exists()


def test_main_uses_config_preset_and_cli_override(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    capsys: pytest.CaptureFixture[str],
) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    (input_dir / "sales.csv").write_text(
        "date,product,category,quantity,unit_price\n2026-04-01,りんご,果物,10,120\n",
        encoding="utf-8",
    )
    config_file = tmp_path / "config.json"
    config_file.write_text(
        json.dumps(
            {
                "input_dir": str(input_dir),
                "output_dir": str(output_dir),
                "presets": {
                    "april": {
                        "month": "2026-04",
                        "output_name": "preset_report.xlsx",
                    }
                },
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(main, "LOG_DIR", tmp_path / "logs")
    monkeypatch.setattr(main, "LOG_FILE", tmp_path / "logs" / "app.log")
    monkeypatch.setattr(
        sys,
        "argv",
        ["main.py", "--config", str(config_file), "--preset", "april", "--output-name", "cli_report.xlsx"],
    )

    exit_code = main.main()
    captured = capsys.readouterr()

    assert exit_code == 0
    assert "処理が完了しました" in captured.out
    assert (output_dir / "cli_report.xlsx").exists()
    assert not (output_dir / "preset_report.xlsx").exists()


def test_main_writes_error_report(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    capsys: pytest.CaptureFixture[str],
) -> None:
    input_dir = tmp_path / "input"
    error_report = tmp_path / "errors.csv"
    input_dir.mkdir()
    (input_dir / "sales_bad.csv").write_text(
        "date,product,quantity,unit_price\n2026/04/01,りんご,bad,-100\n",
        encoding="utf-8",
    )
    monkeypatch.setattr(main, "LOG_DIR", tmp_path / "logs")
    monkeypatch.setattr(main, "LOG_FILE", tmp_path / "logs" / "app.log")
    monkeypatch.setattr(sys, "argv", ["main.py", "--input-dir", str(input_dir), "--error-report", str(error_report)])

    exit_code = main.main()
    captured = capsys.readouterr()

    assert exit_code == 1
    assert "エラー一覧CSV:" in captured.out
    assert "売上データの内容に修正が必要" in captured.out
    assert error_report.exists()
    rows = list(csv.DictReader(error_report.open(encoding="utf-8-sig", newline="")))
    assert "fix" in rows[0]
    assert rows[0]["fix"]


def test_main_error_prints_log_file_path(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    capsys: pytest.CaptureFixture[str],
) -> None:
    log_file = tmp_path / "logs" / "app.log"
    monkeypatch.setattr(main, "LOG_DIR", tmp_path / "logs")
    monkeypatch.setattr(main, "LOG_FILE", log_file)
    monkeypatch.setattr(sys, "argv", ["main.py", "--month", "2026-04", "--input-dir", str(tmp_path / "missing_input")])

    exit_code = main.main()
    captured = capsys.readouterr()

    assert exit_code == 1
    assert "ログファイル:" in captured.out
    assert str(log_file.resolve()) in captured.out
    assert "入力フォルダが見つかりません" in captured.out
    assert "確認してください" in captured.out


def test_format_user_error_message_contains_common_checks() -> None:
    message = main.format_user_error_message(FileNotFoundError("読み込み対象の売上データが見つかりません。"))

    assert "レポート作成中にエラーが発生しました" in message
    assert "入力フォルダに売上データがあるか" in message
    assert "読み込みファイル名パターン" in message


def test_repair_legacy_text_files_updates_json_and_jsonl(tmp_path: Path) -> None:
    gui_state_file = tmp_path / "gui_state.json"
    audit_log_file = tmp_path / "report_history.jsonl"
    alias_preset_file = tmp_path / "column_alias_presets.json"

    gui_state_file.write_text(json.dumps({"label": "險ｭ螳壹ヵ繧｡繧､繝ｫ"}, ensure_ascii=False), encoding="utf-8")
    audit_log_file.write_text(
        json.dumps({"error": "險ｭ螳壹ヵ繧｡繧､繝ｫ"}, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    alias_preset_file.write_text(
        json.dumps({"shop_a": {"note": "繝ｭ繧ｰ縺ｯ縺ｾ縺縺ゅｊ縺ｾ縺帙ｓ"}}, ensure_ascii=False),
        encoding="utf-8",
    )

    repaired = main.repair_legacy_text_files(gui_state_file, audit_log_file, alias_preset_file)

    assert repaired[str(gui_state_file)] is True
    assert repaired[str(audit_log_file)] is True
    assert repaired[str(alias_preset_file)] is True
    assert json.loads(gui_state_file.read_text(encoding="utf-8"))["label"] == "設定ファイル"
    assert json.loads(audit_log_file.read_text(encoding="utf-8").strip())["error"] == "設定ファイル"
    assert json.loads(alias_preset_file.read_text(encoding="utf-8"))["shop_a"]["note"] == "ログはまだありません"
    assert len(list(tmp_path.glob("*.json.*.bak"))) == 2
    assert len(list(tmp_path.glob("*.jsonl.*.bak"))) == 1


def test_repair_legacy_text_files_skips_clean_files(tmp_path: Path) -> None:
    gui_state_file = tmp_path / "gui_state.json"
    gui_state_file.write_text(json.dumps({"label": "設定ファイル"}, ensure_ascii=False), encoding="utf-8")

    repaired = main.repair_legacy_text_files(gui_state_file, tmp_path / "missing.jsonl", tmp_path / "missing.json")

    assert repaired[str(gui_state_file)] is False


def test_main_repair_legacy_text_option(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    capsys: pytest.CaptureFixture[str],
) -> None:
    gui_state_file = tmp_path / "gui_state.json"
    audit_log_file = tmp_path / "report_history.jsonl"
    alias_preset_file = tmp_path / "column_alias_presets.json"
    gui_state_file.write_text(json.dumps({"label": "險ｭ螳壹ヵ繧｡繧､繝ｫ"}, ensure_ascii=False), encoding="utf-8")
    audit_log_file.write_text(json.dumps({"error": "險ｭ螳壹ヵ繧｡繧､繝ｫ"}, ensure_ascii=False) + "\n", encoding="utf-8")
    alias_preset_file.write_text(json.dumps({"shop_a": {"note": "險ｭ螳壹ヵ繧｡繧､繝ｫ"}}, ensure_ascii=False), encoding="utf-8")

    monkeypatch.setattr(main, "LOG_DIR", tmp_path)
    monkeypatch.setattr(main, "AUDIT_LOG_FILE", audit_log_file)
    monkeypatch.setattr(main, "ALIAS_PRESET_FILE", alias_preset_file)
    monkeypatch.setattr(sys, "argv", ["main.py", "--repair-legacy-text"])

    exit_code = main.main()
    captured = capsys.readouterr()

    assert exit_code == 0
    assert "修復しました" in captured.out
    assert "gui_state.json" in captured.out
