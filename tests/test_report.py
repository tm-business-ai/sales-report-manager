from pathlib import Path
import os

import pandas as pd
import pytest
from openpyxl import load_workbook

import report


def make_sales_dataframe() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"date": "2026-04-01", "product": "りんご", "category": "果物", "quantity": "10", "unit_price": "120"},
            {"date": "2026-04-02", "product": "みかん", "category": "果物", "quantity": "8", "unit_price": "100"},
            {"date": "2026-05-01", "product": "りんご", "category": "果物", "quantity": "5", "unit_price": "120"},
        ]
    )


def write_csv(path: Path, rows: list[str], encoding: str = "utf-8") -> None:
    path.write_text("\n".join(rows) + "\n", encoding=encoding)


def test_read_csv_files_merges_files_and_adds_source_file(tmp_path: Path) -> None:
    write_csv(tmp_path / "sales_1.csv", ["date,product,quantity,unit_price", "2026-04-01,りんご,10,120"])
    write_csv(tmp_path / "sales_2.csv", ["date,product,quantity,unit_price", "2026-04-02,みかん,8,100"])

    df = report.read_csv_files(tmp_path)

    assert len(df) == 2
    assert list(df["source_file"]) == ["sales_1.csv", "sales_2.csv"]
    assert list(df["source_row"]) == [2, 2]


def test_read_csv_files_ignores_sample_files_and_supports_pattern(tmp_path: Path) -> None:
    write_csv(tmp_path / "sample_sales.csv", ["date,product,quantity,unit_price", "2026-04-01,サンプル,99,999"])
    write_csv(tmp_path / "sales_target.csv", ["date,product,quantity,unit_price", "2026-04-01,りんご,10,120"])
    write_csv(tmp_path / "other.csv", ["date,product,quantity,unit_price", "2026-04-01,みかん,8,100"])

    df = report.read_csv_files(tmp_path, pattern="sales_*.csv")

    assert len(df) == 1
    assert df.loc[0, "source_file"] == "sales_target.csv"


def test_read_csv_files_supports_cp932_input(tmp_path: Path) -> None:
    write_csv(tmp_path / "sales_cp932.csv", ["date,product,quantity,unit_price", "2026-04-01,りんご,10,120"], encoding="cp932")

    df = report.read_csv_files(tmp_path)

    assert df.loc[0, "product"] == "りんご"


def test_read_csv_files_reports_no_target_csv(tmp_path: Path) -> None:
    write_csv(tmp_path / "sample_sales.csv", ["date,product,quantity,unit_price", "2026-04-01,サンプル,99,999"])

    with pytest.raises(FileNotFoundError, match="読み込み対象の売上データ") as exc_info:
        report.read_csv_files(tmp_path)

    message = str(exc_info.value)
    assert "確認してください" in message
    assert "sales_*.csv" in message


def test_read_sales_files_supports_xlsx_and_pattern_expansion(tmp_path: Path) -> None:
    excel_file = tmp_path / "sales_2026_04.xlsx"
    pd.DataFrame(
        [
            {"date": "2026-04-01", "product": "apple", "category": "fruit", "quantity": 10, "unit_price": 120},
            {"date": "2026-05-01", "product": "orange", "category": "fruit", "quantity": 8, "unit_price": 100},
        ]
    ).to_excel(excel_file, index=False)

    df = report.read_sales_files(tmp_path, pattern="sales_*.csv")
    validated_df = report.validate_data(df)
    filtered_df = report.filter_data(validated_df, month="2026-04")
    summary_df = report.create_summary(filtered_df, "product")

    assert list(df["source_file"]) == ["sales_2026_04.xlsx", "sales_2026_04.xlsx"]
    assert len(filtered_df) == 1
    assert summary_df.loc[0, "total_amount"] == 1200


def test_read_sales_files_supports_xlsx_column_aliases(tmp_path: Path) -> None:
    excel_file = tmp_path / "sales_alias.xlsx"
    pd.DataFrame(
        [{"sold_at": "2026-04-01", "item": "apple", "dept": "fruit", "qty": 10, "price": 120}]
    ).to_excel(excel_file, index=False)

    df = report.read_sales_files(excel_file)
    validated_df = report.validate_data(
        df,
        {
            "sold_at": "date",
            "item": "product",
            "dept": "category",
            "qty": "quantity",
            "price": "unit_price",
        },
    )

    assert validated_df.loc[0, "amount"] == 1200
    assert validated_df.loc[0, "category"] == "fruit"


def test_read_sales_columns_reads_first_input_file_columns(tmp_path: Path) -> None:
    write_csv(tmp_path / "sales_2026_04.csv", ["売上日,品名,商品分類,販売数,販売単価", "2026-04-01,apple,fruit,10,120"])

    columns = report.read_sales_columns(tmp_path, pattern="sales_*.csv")

    assert columns == ("売上日", "品名", "商品分類", "販売数", "販売単価")


def test_infer_column_aliases_supports_common_japanese_names() -> None:
    aliases = report.infer_column_aliases(("売上日", "品名", "商品分類", "販売数", "販売単価", "売上金額"))

    assert aliases == {
        "売上日": "date",
        "品名": "product",
        "商品分類": "category",
        "販売数": "quantity",
        "販売単価": "unit_price",
        "売上金額": "amount",
    }


def test_missing_required_column_labels_uses_aliases() -> None:
    columns = ("売上日", "品名", "販売数")
    aliases = report.infer_column_aliases(columns)

    missing = report.missing_required_column_labels(columns, aliases)

    assert missing == ["単価"]


def test_read_sales_files_rejects_explicit_unsupported_file(tmp_path: Path) -> None:
    text_file = tmp_path / "sales.txt"
    text_file.write_text("date,product,quantity,unit_price\n", encoding="utf-8")

    with pytest.raises(ValueError, match="対応していない入力ファイル形式"):
        report.read_sales_files(text_file)


def test_read_sales_files_ignores_unsupported_files_in_directory(tmp_path: Path) -> None:
    write_csv(tmp_path / "sales.csv", ["date,product,quantity,unit_price", "2026-04-01,apple,10,120"])
    (tmp_path / "memo.txt").write_text("ignore me", encoding="utf-8")

    df = report.read_sales_files(tmp_path)

    assert len(df) == 1
    assert df.loc[0, "source_file"] == "sales.csv"


def test_read_xls_reports_xlrd_requirement_when_missing(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    xls_file = tmp_path / "sales.xls"
    xls_file.write_bytes(b"not a real xls")
    monkeypatch.setattr(report.importlib.util, "find_spec", lambda name: None if name == "xlrd" else object())

    with pytest.raises(ImportError, match="xlrd"):
        report.read_sales_files(xls_file)


def test_validate_data_accepts_japanese_column_names() -> None:
    df = pd.DataFrame([{"売上日": "2026-04-01", "商品名": "りんご", "カテゴリ": "果物", "数量": "10", "単価": "120"}])

    validated_df = report.validate_data(df)

    assert list(validated_df.columns) == ["date", "product", "category", "quantity", "unit_price", "amount"]
    assert validated_df.loc[0, "amount"] == 1200


def test_validate_data_accepts_custom_column_aliases() -> None:
    df = pd.DataFrame(
        [
            {
                "sold_at": "2026-04-01",
                "item": "apple",
                "dept": "fruit",
                "qty": "10",
                "price": "120",
            }
        ]
    )

    validated_df = report.validate_data(
        df,
        {
            "sold_at": "date",
            "item": "product",
            "dept": "category",
            "qty": "quantity",
            "price": "unit_price",
        },
    )

    assert list(validated_df.columns) == ["date", "product", "category", "quantity", "unit_price", "amount"]
    assert validated_df.loc[0, "amount"] == 1200


def test_validate_data_missing_columns_has_user_friendly_guidance() -> None:
    df = pd.DataFrame([{"数量": "1", "単価": "100"}])

    with pytest.raises(report.DataValidationError) as exc_info:
        report.validate_data(df)

    message = str(exc_info.value)
    assert "売上データに必要な列が不足" in message
    assert "不足している列" in message
    assert "日付" in message
    assert "商品名" in message
    assert "必要な列の例" in message
    assert "修正方法" in message


def test_validate_data_collects_multiple_errors_and_writes_report(tmp_path: Path) -> None:
    df = pd.DataFrame(
        [
            {"date": "2026/99/99", "product": "りんご", "quantity": "bad", "unit_price": "-120"},
            {"date": "2026-04-02", "product": "みかん", "quantity": "-1", "unit_price": "bad"},
        ]
    )

    with pytest.raises(report.DataValidationError) as exc_info:
        report.validate_data(df)

    message = str(exc_info.value)
    assert "日付として読み取れない値" in message
    assert "2026/99/99" in message
    assert "数量に数値以外の値" in message
    assert "単価に数値以外の値" in message
    assert "数量にマイナス" in message
    assert "単価にマイナス" in message
    assert "修正方法" in message

    output_file = report.write_validation_error_report(exc_info.value, tmp_path / "errors.csv")
    error_df = pd.read_csv(output_file)
    assert "fix" in error_df.columns
    assert any("数量には" in value for value in error_df["fix"])
    assert set(error_df["issue"]) == {"invalid_date", "invalid_quantity", "negative_unit_price", "negative_quantity", "invalid_unit_price"}


def test_validate_data_reports_invalid_amount_when_amount_column_exists() -> None:
    df = pd.DataFrame([{"日付": "2026-04-01", "商品名": "りんご", "数量": "1", "単価": "100", "金額": "abc"}])

    with pytest.raises(report.DataValidationError) as exc_info:
        report.validate_data(df)

    assert "金額に数値以外の値" in str(exc_info.value)
    assert "金額列を使う場合" in str(exc_info.value)


def test_filter_data_supports_date_range_product_and_category() -> None:
    df = report.validate_data(make_sales_dataframe())

    filtered = report.filter_data(df, start_date="2026-04-01", end_date="2026-04-30", product="りんご", category="果物")

    assert len(filtered) == 1
    assert filtered.iloc[0]["product"] == "りんご"


def test_filter_data_rejects_month_and_date_range_together() -> None:
    df = report.validate_data(make_sales_dataframe())

    with pytest.raises(ValueError, match="同時に指定できません"):
        report.filter_data(df, month="2026-04", start_date="2026-04-01")


def test_filter_data_rejects_start_date_after_end_date() -> None:
    df = report.validate_data(make_sales_dataframe())

    with pytest.raises(ValueError, match="開始日が終了日より後"):
        report.filter_data(df, start_date="2026-04-30", end_date="2026-04-01")


def test_filter_data_no_month_data_has_confirmation_points() -> None:
    df = report.validate_data(make_sales_dataframe())

    with pytest.raises(ValueError) as exc_info:
        report.filter_data(df, month="2026-06")

    message = str(exc_info.value)
    assert "対象月に一致する売上データがありません" in message
    assert "確認してください" in message
    assert "対象月: 2026-06" in message


def test_create_summary_monthly_trend_and_summary_csv(tmp_path: Path) -> None:
    df = report.validate_data(make_sales_dataframe())
    summaries = report.create_summaries(df, "product", all_summaries=True)
    trend = report.create_monthly_trend(df)
    daily = report.create_daily_trend(df)
    csv_files = report.write_summary_csvs(summaries, tmp_path, "sales")

    assert list(summaries["product"].columns) == [
        "rank",
        "product",
        "total_amount",
        "transaction_count",
        "total_quantity",
        "average_unit_price",
        "total_amount_ratio",
    ]
    assert summaries["product"].loc[0, "rank"] == 1
    assert summaries["product"].loc[0, "transaction_count"] == 2
    assert summaries["product"].loc[0, "average_unit_price"] == 120
    assert list(trend["month"]) == ["2026-04", "2026-05"]
    assert list(trend["total_amount"]) == [2000, 600]
    assert list(daily["date"]) == ["2026-04-01", "2026-04-02", "2026-05-01"]
    assert list(daily["transaction_count"]) == [1, 1, 1]
    assert {path.name for path in csv_files} == {"sales_product.csv", "sales_category.csv"}
    assert pd.read_csv(tmp_path / "sales_product.csv").loc[0, "商品"] == "りんご"


def test_create_summary_and_daily_trend_handle_zero_quantity() -> None:
    df = report.validate_data(
        pd.DataFrame(
            [
                {"date": "2026-04-01", "product": "sample", "category": "test", "quantity": "0", "unit_price": "120"},
            ]
        )
    )

    summary = report.create_summary(df, "product")
    daily = report.create_daily_trend(df)

    assert summary.loc[0, "average_unit_price"] == 0
    assert daily.loc[0, "average_unit_price"] == 0


def test_create_category_summary_includes_rank_count_and_average() -> None:
    df = report.validate_data(
        pd.DataFrame(
            [
                {"date": "2026-04-01", "product": "a", "category": "fruit", "quantity": "10", "unit_price": "100"},
                {"date": "2026-04-02", "product": "b", "category": "fruit", "quantity": "5", "unit_price": "200"},
                {"date": "2026-04-03", "product": "c", "category": "drink", "quantity": "2", "unit_price": "100"},
            ]
        )
    )

    summary = report.create_summary(df, "category")

    assert list(summary["rank"]) == [1, 2]
    assert summary.loc[0, "category"] == "fruit"
    assert summary.loc[0, "transaction_count"] == 2
    assert summary.loc[0, "total_quantity"] == 15
    assert summary.loc[0, "average_unit_price"] == 2000 / 15


def test_create_month_end_summary_includes_key_metrics_and_top_items() -> None:
    df = report.validate_data(
        pd.DataFrame(
            [
                {"date": "2026-04-01", "product": "apple", "category": "fruit", "quantity": "10", "unit_price": "100"},
                {"date": "2026-04-02", "product": "orange", "category": "fruit", "quantity": "5", "unit_price": "200"},
                {"date": "2026-04-03", "product": "notebook", "category": "stationery", "quantity": "2", "unit_price": "300"},
            ]
        )
    )

    summary = report.create_month_end_summary(df, "2026-04", uncategorized_count=2, error_count=1)
    values = dict(summary.to_records(index=False))

    assert values["対象月"] == "2026-04"
    assert values["集計期間"] == "2026-04-01 ～ 2026-04-30"
    assert values["売上合計"] == 2600
    assert values["明細件数"] == 3
    assert values["数量合計"] == 17
    assert values["平均単価"] == 2600 / 17
    assert values["対象日数"] == 3
    assert values["商品数"] == 3
    assert values["カテゴリ数"] == 2
    assert values["売上トップ商品"] == "apple"
    assert values["売上トップ商品の売上"] == 1000
    assert values["売上トップカテゴリ"] == "fruit"
    assert values["売上トップカテゴリの売上"] == 2000
    assert values["未分類データ件数"] == 2
    assert values["エラー行件数"] == 1
    assert values["確認が必要な件数"] == 3


def test_create_previous_month_comparison_calculates_difference_and_rate() -> None:
    current_df = report.validate_data(
        pd.DataFrame(
            [
                {"date": "2026-04-01", "product": "apple", "category": "fruit", "quantity": "10", "unit_price": "100"},
                {"date": "2026-04-02", "product": "orange", "category": "fruit", "quantity": "5", "unit_price": "200"},
            ]
        )
    )
    previous_df = report.validate_data(
        pd.DataFrame(
            [
                {"date": "2026-03-01", "product": "apple", "category": "fruit", "quantity": "5", "unit_price": "100"},
            ]
        )
    )

    comparison = report.create_previous_month_comparison(current_df, previous_df, "2026-04")
    amount_row = comparison[comparison["指標"] == "売上合計"].iloc[0]

    assert amount_row["当月"] == 2000
    assert amount_row["前月"] == 500
    assert amount_row["差分"] == 1500
    assert amount_row["増減率"] == 3


def test_create_previous_month_comparison_handles_missing_previous_data() -> None:
    current_df = report.validate_data(
        pd.DataFrame(
            [
                {"date": "2026-04-01", "product": "apple", "category": "fruit", "quantity": "10", "unit_price": "100"},
            ]
        )
    )
    empty_previous_df = current_df.iloc[0:0].copy()

    comparison = report.create_previous_month_comparison(current_df, empty_previous_df, "2026-04")
    amount_row = comparison[comparison["指標"] == "売上合計"].iloc[0]

    assert amount_row["前月"] == 0
    assert amount_row["増減率"] == "比較不可"
    assert "前月データなし" in set(comparison["前月"])


def test_create_uncategorized_rows_extracts_review_targets() -> None:
    df = report.validate_data(
        pd.DataFrame(
            [
                {"date": "2026-04-01", "product": "", "category": "果物", "quantity": "1", "unit_price": "100"},
                {"date": "2026-04-02", "product": "りんご", "category": "", "quantity": "1", "unit_price": "100"},
                {"date": "2026-04-03", "product": "みかん", "category": "未分類", "quantity": "1", "unit_price": "100"},
                {"date": "2026-04-04", "product": "バナナ", "category": "その他", "quantity": "1", "unit_price": "100"},
                {"date": "2026-04-05", "product": "未設定", "category": "果物", "quantity": "1", "unit_price": "100"},
                {"date": "2026-04-06", "product": "ノート", "category": "文具", "quantity": "1", "unit_price": "100"},
            ]
        )
    )
    df["source_file"] = "sales.csv"
    df["source_row"] = range(2, 8)

    uncategorized = report.create_uncategorized_rows(df)

    assert len(uncategorized) == 5
    assert "商品名が空欄" in set(uncategorized["reason"])
    assert "カテゴリが空欄" in set(uncategorized["reason"])
    assert "カテゴリが未分類" in set(uncategorized["reason"])
    assert "カテゴリがその他" in set(uncategorized["reason"])
    assert "商品名が未設定" in set(uncategorized["reason"])
    assert set(uncategorized["source_file"]) == {"sales.csv"}


def test_create_uncategorized_rows_returns_empty_dataframe_for_clean_data() -> None:
    df = report.validate_data(
        pd.DataFrame(
            [{"date": "2026-04-01", "product": "りんご", "category": "果物", "quantity": "1", "unit_price": "100"}]
        )
    )

    uncategorized = report.create_uncategorized_rows(df)

    assert uncategorized.empty


def test_create_summary_rejects_missing_group_by_column() -> None:
    df = report.validate_data(pd.DataFrame([{"date": "2026-04-01", "product": "りんご", "quantity": "10", "unit_price": "120"}]))

    with pytest.raises(ValueError, match="必要な列"):
        report.create_summary(df, group_by="category")


def test_prepare_detail_for_excel_ignores_input_amount_column() -> None:
    df = report.validate_data(
        pd.DataFrame(
            [
                {
                    "日付": "2026-04-01",
                    "商品名": "オーガニックコーヒー",
                    "カテゴリ": "飲料",
                    "数量": "2",
                    "単価": "780",
                    "金額": "999999",
                }
            ]
        )
    )

    excel_df = report.prepare_detail_for_excel(df)

    assert list(excel_df.columns).count("金額") == 1
    assert excel_df.loc[0, "金額"] == 1560
    assert excel_df.loc[1, "金額"] == 1560


def test_save_to_excel_writes_charts_custom_name_and_conditions(tmp_path: Path) -> None:
    detail_df = report.filter_data(report.validate_data(make_sales_dataframe()), start_date="2026-04-01", end_date="2026-05-31")
    summaries = report.create_summaries(detail_df, "product", all_summaries=True)
    trend = report.create_monthly_trend(detail_df)
    daily = report.create_daily_trend(detail_df)
    uncategorized = report.create_uncategorized_rows(detail_df)
    errors = report.create_validation_error_rows()
    month_end = report.create_month_end_summary(detail_df, "2026-04", uncategorized_count=len(uncategorized), error_count=len(errors))
    comparison = report.create_previous_month_comparison(detail_df, detail_df.iloc[0:0].copy(), "2026-04")

    output_file = report.save_to_excel(
        detail_df,
        summaries,
        tmp_path,
        None,
        "product",
        input_dir=tmp_path / "input",
        pattern="sales_*.csv",
        all_summaries=True,
        start_date="2026-04-01",
        end_date="2026-05-31",
        product="りんご",
        category="果物",
        daily_trend_df=daily,
        monthly_trend_df=trend,
        month_end_summary_df=month_end,
        previous_month_comparison_df=comparison,
        uncategorized_df=uncategorized,
        validation_error_df=errors,
        output_name="custom_report",
        charts=True,
        summary_csv_dir=tmp_path / "csv",
    )

    workbook = load_workbook(output_file, data_only=True)

    assert output_file.name == "custom_report.xlsx"
    assert workbook.sheetnames == ["月末サマリー", "前月比較", "詳細データ", "商品別集計", "カテゴリ別集計", "日別推移", "月別推移", "未分類データ", "エラー行一覧", "実行条件"]
    assert workbook["詳細データ"]["A4"].value == "日付"
    assert workbook["商品別集計"]["A4"].value == "順位"
    assert workbook["商品別集計"]["D4"].value == "件数"
    assert workbook["商品別集計"]["F4"].value == "平均単価"
    assert workbook["日別推移"]["A4"].value == "日付"
    assert workbook["日別推移"]["B4"].value == "売上合計"
    assert workbook["月別推移"]["A4"].value == "月"
    assert workbook["月末サマリー"]["A3"].value == "項目"
    assert workbook["前月比較"]["A3"].value == "指標"
    assert workbook["未分類データ"]["A3"].value == "メッセージ"
    assert workbook["未分類データ"]["A4"].value == "確認が必要な未分類データはありません"
    assert workbook["エラー行一覧"]["A3"].value == "メッセージ"
    assert workbook["エラー行一覧"]["A4"].value == "検証エラーはありません"
    assert workbook["実行条件"]["B6"].value == "2026-04-01"
    assert workbook["実行条件"]["B7"].value == "2026-05-31"
    assert workbook["実行条件"]["B8"].value == "りんご"
    assert workbook["実行条件"]["B9"].value == "果物"
    assert workbook["実行条件"]["B15"].value == "custom_report"
    assert workbook["実行条件"]["B16"].value == str(tmp_path / "csv")
    assert len(workbook["商品別集計"]._charts) == 1
    assert len(workbook["日別推移"]._charts) == 1
    assert len(workbook["月別推移"]._charts) == 1


def test_prepare_validation_errors_for_excel_includes_fix_column() -> None:
    error = report.DataValidationError(
        [
            report.ValidationIssue(
                "invalid_quantity",
                "数量に数値以外の値があります。不正な値: abc",
                "sales.csv",
                2,
                "数量には 1、2、10 のような0以上の数値を入力してください。",
            )
        ]
    )

    excel_df = report.prepare_validation_errors_for_excel(report.create_validation_error_rows(error))

    assert "修正方法" in excel_df.columns
    assert excel_df.loc[0, "エラー内容"].startswith("数量に数値以外")
    assert "数量には" in excel_df.loc[0, "修正方法"]


def test_create_validation_error_rows_includes_source_values() -> None:
    error = report.DataValidationError(
        [
            report.ValidationIssue(
                "invalid_quantity",
                "数量に数値以外の値があります。不正な値: abc",
                "sales.csv",
                2,
                "数量には 1、2、10 のような0以上の数値を入力してください。",
                {"date": "2026-04-01", "product": "りんご", "quantity": "abc", "unit_price": "100"},
            )
        ]
    )

    rows = report.create_validation_error_rows(error)

    assert rows.loc[0, "fix"].startswith("数量には")
    assert rows.loc[0, "date"] == "2026-04-01"
    assert rows.loc[0, "product"] == "りんご"
    assert rows.loc[0, "quantity"] == "abc"


def test_save_to_excel_writes_validation_error_fix_column(tmp_path: Path) -> None:
    detail_df = report.filter_data(report.validate_data(make_sales_dataframe()), month="2026-04")
    summaries = report.create_summaries(detail_df, "product", all_summaries=False)
    error = report.DataValidationError(
        [
            report.ValidationIssue(
                "invalid_date",
                "日付として読み取れない値があります。不正な値: 2026/99/99",
                "sales.csv",
                2,
                "日付列を 2026-04-01 または 2026/04/01 のような形式に修正してください。",
            )
        ]
    )

    output_file = report.save_to_excel(
        detail_df,
        summaries,
        tmp_path,
        "2026-04",
        "product",
        input_dir=tmp_path / "input",
        pattern="sales*.csv",
        validation_error_df=report.create_validation_error_rows(error),
    )

    workbook = load_workbook(output_file, data_only=True)
    sheet = workbook["エラー行一覧"]
    headers = [sheet.cell(row=3, column=column).value for column in range(1, 6)]

    assert "修正方法" in headers
    assert sheet.cell(row=4, column=headers.index("修正方法") + 1).value.startswith("日付列を")


def test_cleanup_old_reports_keeps_newest_reports(tmp_path: Path) -> None:
    old_report = tmp_path / "monthly_report_202601_20260425_100000.xlsx"
    middle_report = tmp_path / "monthly_report_202602_20260425_100000.xlsx"
    new_report = tmp_path / "monthly_report_202603_20260425_100000.xlsx"
    unrelated_file = tmp_path / "memo.xlsx"

    for index, file in enumerate([old_report, middle_report, new_report, unrelated_file], start=1):
        file.write_text("dummy", encoding="utf-8")
        os.utime(file, (index, index))

    deleted_reports = report.cleanup_old_reports(tmp_path, keep_reports=2)

    assert deleted_reports == [old_report]
    assert not old_report.exists()
    assert middle_report.exists()
    assert new_report.exists()
    assert unrelated_file.exists()
